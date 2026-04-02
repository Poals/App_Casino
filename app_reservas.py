import streamlit as st
import sqlite3
from datetime import datetime, timedelta
import pandas as pd
import io

# ====================== BASE DE DATOS ======================


@st.cache_resource
def init_conexion():
    """Inicializa la conexión a la base de datos"""
    conn = sqlite3.connect('reservas_comidas.db', check_same_thread=False)
    return conn


conn = init_conexion()
c = conn.cursor()

# Crear tablas si no existen
c.execute('''
CREATE TABLE IF NOT EXISTS usuarios (
    id INTEGER PRIMARY KEY,
    nombre TEXT UNIQUE,
    password TEXT,
    rol TEXT,
    nombre_completo TEXT
)
''')

# Si la base ya existía sin nombre_completo, se intenta agregar
try:
    c.execute("ALTER TABLE usuarios ADD COLUMN nombre_completo TEXT")
except sqlite3.OperationalError:
    pass

c.execute('''
CREATE TABLE IF NOT EXISTS reservas (
    id INTEGER PRIMARY KEY,
    fecha_comida TEXT,      -- fecha para la que se reserva (YYYY-MM-DD)
    nombre_usuario TEXT,
    desayuno INTEGER DEFAULT 0,
    almuerzo INTEGER DEFAULT 0,
    cena INTEGER DEFAULT 0,
    reservado_en TEXT
)
''')
conn.commit()

# Usuarios de prueba (puedes agregar más después)
c.execute("INSERT OR IGNORE INTO usuarios (nombre, password, rol, nombre_completo) VALUES (?, ?, ?, ?)",
          ("jhon", "123", "empleado", "Jhon Nombre"))
c.execute("INSERT OR IGNORE INTO usuarios (nombre, password, rol, nombre_completo) VALUES (?, ?, ?, ?)",
          ("maria", "123", "empleado", "Maria Nombre"))
c.execute("INSERT OR IGNORE INTO usuarios (nombre, password, rol, nombre_completo) VALUES (?, ?, ?, ?)",
          ("admin", "123", "admin", "Administrador"))
# Forzar contraseña del admin a 123 para evitar que base vieja mantenga contraseña previa
c.execute("UPDATE usuarios SET password=?, rol=?, nombre_completo=? WHERE nombre=?",
          ("123", "admin", "Administrador", "admin"))
conn.commit()

# ====================== FUNCIONES ======================


def puede_reservar_hoy():
    """Devuelve True si aún se puede reservar para mañana (antes de las 22:00)"""
    ahora = datetime.now()
    hora_limite = datetime(ahora.year, ahora.month, ahora.day, 22, 0, 0)
    return ahora < hora_limite


def obtener_manana():
    return (datetime.now() + timedelta(days=1)).date().isoformat()


def obtener_reserva_hoy(usuario, fecha):
    """Obtiene la reserva existente del usuario para esa fecha"""
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, desayuno, almuerzo, cena FROM reservas WHERE nombre_usuario=? AND fecha_comida=?", (usuario, fecha))
        return cursor.fetchone()
    except Exception as e:
        st.error(f"Error al obtener reserva: {e}")
        return None


def validar_cedula(cedula):
    """Valida que la cédula sea solo números"""
    return cedula.isdigit()


def usuario_existe(cedula):
    """Verifica si el usuario (cédula) ya existe"""
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM usuarios WHERE nombre=?", (cedula,))
        return cursor.fetchone() is not None
    except Exception as e:
        st.error(f"Error al verificar usuario: {e}")
        return False


def registrar_usuario(nombre_completo, cedula):
    """Registra un nuevo usuario con cédula como usuario y últimos 4 dígitos como contraseña"""
    try:
        if not validar_cedula(cedula):
            st.error("❌ La cédula debe contener solo números")
            return False

        if len(cedula) < 4:
            st.error("❌ La cédula debe tener al menos 4 dígitos")
            return False

        if usuario_existe(cedula):
            st.error("❌ Esta cédula ya está registrada")
            return False

        # Generar contraseña con los últimos 4 dígitos
        contraseña = cedula[-4:]

        # Insertar usuario
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO usuarios (nombre, password, rol, nombre_completo) VALUES (?, ?, ?, ?)",
            (cedula, contraseña, "empleado", nombre_completo)
        )
        conn.commit()

        return True, cedula, contraseña
    except Exception as e:
        st.error(f"Error al registrar usuario: {e}")
        return False


def mostrar_pagina_empleado():
    """Página principal para empleados"""
    usuario = st.session_state.usuario
    nombre_completo = st.session_state.get("nombre_completo", usuario)

    st.sidebar.success(f"Conectado como: **{nombre_completo}** (Empleado)")
    st.sidebar.write(f"Cédula: {usuario}")
    st.sidebar.write(f"Nombre: {nombre_completo}")

    if st.sidebar.button("Cerrar sesión"):
        del st.session_state.usuario
        del st.session_state.rol
        if "nombre_completo" in st.session_state:
            del st.session_state.nombre_completo
        st.rerun()

    fecha_manana = obtener_manana()
    hora_actual = datetime.now().strftime("%Y/%m/%d %H:%M")
    st.write(f"🕒 Fecha y hora actual: {hora_actual}")
    st.subheader(f"📅 Reserva para mañana ({fecha_manana})")

    if puede_reservar_hoy():
        st.success("✅ Aún estás a tiempo de reservar")

        # Verificar si ya existe reserva para hoy
        reserva_actual = obtener_reserva_hoy(usuario, fecha_manana)

        if reserva_actual:
            st.warning(
                "⚠️ Ya tienes una reserva para mañana. Puedes actualizarla aquí.")
            id_reserva, desayuno_actual, almuerzo_actual, cena_actual = reserva_actual
            desayuno = st.checkbox(
                "🌅 Desayuno", value=bool(desayuno_actual))
            almuerzo = st.checkbox(
                "🍽️ Almuerzo", value=bool(almuerzo_actual))
            cena = st.checkbox("🌙 Cena", value=bool(cena_actual))

            col1, col2 = st.columns(2)
            with col1:
                if st.button("Actualizar reserva", type="primary"):
                    c.execute('''
                        UPDATE reservas 
                        SET desayuno=?, almuerzo=?, cena=?, reservado_en=?
                        WHERE id=?
                    ''', (int(desayuno), int(almuerzo), int(cena), datetime.now().isoformat(), id_reserva))
                    conn.commit()
                    st.success("✅ Reserva actualizada correctamente!")
                    st.rerun()
            with col2:
                if st.button("Cancelar reserva", type="secondary"):
                    c.execute("DELETE FROM reservas WHERE id=?",
                              (id_reserva,))
                    conn.commit()
                    st.success("✅ Reserva cancelada")
                    st.rerun()
        else:
            desayuno = st.checkbox("🌅 Desayuno", value=True)
            almuerzo = st.checkbox("🍽️ Almuerzo", value=True)
            cena = st.checkbox("🌙 Cena", value=False)

            if st.button("Guardar mi reserva", type="primary"):
                c.execute('''
                    INSERT INTO reservas 
                    (fecha_comida, nombre_usuario, desayuno, almuerzo, cena, reservado_en)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (fecha_manana, usuario, int(desayuno), int(almuerzo), int(cena), datetime.now().isoformat()))
                conn.commit()
                st.success("✅ Reserva guardada correctamente!")
                st.rerun()
    else:
        st.error(
            "⛔ Ya pasó la hora límite (22:00). No se pueden hacer más reservas para mañana.")

    # Mis reservas anteriores
    st.subheader("📋 Mis reservas anteriores")
    c.execute("SELECT fecha_comida, desayuno, almuerzo, cena, reservado_en FROM reservas WHERE nombre_usuario=? ORDER BY fecha_comida DESC", (usuario,))
    mis_reservas = c.fetchall()
    if mis_reservas:
        df_mis = pd.DataFrame(mis_reservas, columns=[
                              "Fecha", "Desayuno", "Almuerzo", "Cena", "Registrado En"])
        df_mis["Desayuno"] = df_mis["Desayuno"].map({1: "✅", 0: ""})
        df_mis["Almuerzo"] = df_mis["Almuerzo"].map({1: "✅", 0: ""})
        df_mis["Cena"] = df_mis["Cena"].map({1: "✅", 0: ""})

        # Formato amigable para la columna de registro
        df_mis["Registrado En"] = pd.to_datetime(
            df_mis["Registrado En"]).dt.strftime("%Y/%m/%d %H:%M")

        st.dataframe(df_mis, width='stretch')
    else:
        st.info("Aún no tienes reservas.")


def mostrar_pagina_admin():
    """Página especial para administradores"""
    usuario = st.session_state.usuario

    st.sidebar.success(f"Conectado como: **{usuario}** (Administrador)")
    if st.sidebar.button("Cerrar sesión"):
        del st.session_state.usuario
        del st.session_state.rol
        st.rerun()

    st.subheader("👑 Panel de Administración")

    # Filtros de fecha
    st.write("### 📅 Filtros de búsqueda")
    col1, col2, col3 = st.columns(3)

    with col1:
        fecha_inicio = st.date_input(
            "Fecha inicio",
            value=datetime.now().replace(day=1),  # Primer día del mes actual
            key="fecha_inicio"
        )

    with col2:
        fecha_fin = st.date_input(
            "Fecha fin",
            value=datetime.now() + timedelta(days=30),  # Un mes adelante
            key="fecha_fin"
        )

    with col3:
        if st.button("🔄 Limpiar filtros", key="limpiar_filtros"):
            st.rerun()

    # Estadísticas en tiempo real
    st.write("### 📊 Estadísticas en tiempo real")
    estadisticas = obtener_estadisticas(
        fecha_inicio.isoformat(), fecha_fin.isoformat())

    if estadisticas:
        df_stats = pd.DataFrame(estadisticas, columns=[
                                "Fecha", "Desayunos", "Almuerzos", "Cenas", "Total Empleados"])

        # Métricas principales
        total_desayunos = df_stats["Desayunos"].sum()
        total_almuerzos = df_stats["Almuerzos"].sum()
        total_cenas = df_stats["Cenas"].sum()
        total_empleados = df_stats["Total Empleados"].max(
        ) if not df_stats.empty else 0

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("🌅 Desayunos", f"{total_desayunos}")
        with col2:
            st.metric("🍽️ Almuerzos", f"{total_almuerzos}")
        with col3:
            st.metric("🌙 Cenas", f"{total_cenas}")
        with col4:
            st.metric("👥 Empleados", f"{total_empleados}")

        st.dataframe(df_stats, width='stretch')

        # Gráfico
        df_grafico = df_stats.set_index(
            "Fecha")[["Desayunos", "Almuerzos", "Cenas"]].astype(int)
        st.bar_chart(df_grafico)
    else:
        st.info("No hay reservas en el período seleccionado.")

    # Todas las reservas filtradas
    st.write("### 📋 Reservas detalladas")
    reservas_filtradas = obtener_reservas_filtradas(
        fecha_inicio.isoformat(), fecha_fin.isoformat())

    if reservas_filtradas:
        df_reservas = pd.DataFrame(reservas_filtradas, columns=[
            "Fecha", "Cédula", "Desayuno", "Almuerzo", "Cena", "Reservado En", "Nombre Completo"
        ])

        # Convertir booleanos para display
        df_display = df_reservas.copy()
        df_display["Desayuno"] = df_display["Desayuno"].map({1: "✅", 0: ""})
        df_display["Almuerzo"] = df_display["Almuerzo"].map({1: "✅", 0: ""})
        df_display["Cena"] = df_display["Cena"].map({1: "✅", 0: ""})

        # Formato amigable de fecha/hora de reserva
        df_display["Reservado En"] = pd.to_datetime(
            df_display["Reservado En"]).dt.strftime("%Y/%m/%d %H:%M")

        st.dataframe(df_display[["Fecha", "Cédula", "Nombre Completo", "Desayuno", "Almuerzo", "Cena", "Reservado En"]],
                     width='stretch', hide_index=True)

        # Botón de descarga Excel
        excel_buffer = generar_excel_reservas(
            fecha_inicio.isoformat(), fecha_fin.isoformat())
        if excel_buffer:
            st.download_button(
                label="📥 Descargar Excel",
                data=excel_buffer,
                file_name=f"reservas_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_excel"
            )
    else:
        st.info("No hay reservas en el período seleccionado.")


def obtener_reservas_filtradas(fecha_inicio=None, fecha_fin=None):
    """Obtiene reservas filtradas por rango de fechas"""
    try:
        cursor = conn.cursor()
        if fecha_inicio and fecha_fin:
            cursor.execute("""
                SELECT r.fecha_comida, r.nombre_usuario, r.desayuno, r.almuerzo, r.cena, r.reservado_en,
                       COALESCE(u.nombre_completo, u.nombre) as nombre_completo
                FROM reservas r
                LEFT JOIN usuarios u ON r.nombre_usuario = u.nombre
                WHERE r.fecha_comida BETWEEN ? AND ?
                ORDER BY r.fecha_comida DESC, r.nombre_usuario
            """, (fecha_inicio, fecha_fin))
        else:
            cursor.execute("""
                SELECT r.fecha_comida, r.nombre_usuario, r.desayuno, r.almuerzo, r.cena, r.reservado_en,
                       COALESCE(u.nombre_completo, u.nombre) as nombre_completo
                FROM reservas r
                LEFT JOIN usuarios u ON r.nombre_usuario = u.nombre
                ORDER BY r.fecha_comida DESC, r.nombre_usuario
            """)
        return cursor.fetchall()
    except Exception as e:
        st.error(f"Error al obtener reservas filtradas: {e}")
        return []


def obtener_estadisticas(fecha_inicio=None, fecha_fin=None):
    """Obtiene estadísticas de reservas por fecha"""
    try:
        cursor = conn.cursor()
        if fecha_inicio and fecha_fin:
            cursor.execute("""
                SELECT fecha_comida,
                       SUM(desayuno) as desayunos,
                       SUM(almuerzo) as almuerzos,
                       SUM(cena) as cenas,
                       COUNT(DISTINCT nombre_usuario) as total_empleados
                FROM reservas
                WHERE fecha_comida BETWEEN ? AND ?
                GROUP BY fecha_comida
                ORDER BY fecha_comida DESC
            """, (fecha_inicio, fecha_fin))
        else:
            cursor.execute("""
                SELECT fecha_comida,
                       SUM(desayuno) as desayunos,
                       SUM(almuerzo) as almuerzos,
                       SUM(cena) as cenas,
                       COUNT(DISTINCT nombre_usuario) as total_empleados
                FROM reservas
                GROUP BY fecha_comida
                ORDER BY fecha_comida DESC
            """)
        return cursor.fetchall()
    except Exception as e:
        st.error(f"Error al obtener estadísticas: {e}")
        return []


def generar_excel_reservas(fecha_inicio=None, fecha_fin=None):
    """Genera archivo Excel dinámico con gráficos y formato avanzado"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, PieChart, Reference, Series
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter

        reservas = obtener_reservas_filtradas(fecha_inicio, fecha_fin)
        if not reservas:
            return None

        # Crear DataFrame
        df = pd.DataFrame(reservas, columns=[
            "Fecha Comida", "Cédula", "Desayuno", "Almuerzo", "Cena", "Reservado En", "Nombre Completo"
        ])

        # Convertir valores booleanos
        df["Desayuno"] = df["Desayuno"].map({1: "Sí", 0: "No"})
        df["Almuerzo"] = df["Almuerzo"].map({1: "Sí", 0: "No"})
        df["Cena"] = df["Cena"].map({1: "Sí", 0: "No"})

        # Formatear fecha de reserva en formato amigable
        df["Reservado En"] = pd.to_datetime(
            df["Reservado En"]).dt.strftime("%Y/%m/%d %H:%M")

        # Reordenar columnas para que coincidan con el orden visual/exportado
        df = df[["Fecha Comida", "Cédula", "Nombre Completo",
                 "Desayuno", "Almuerzo", "Cena", "Reservado En"]]

        # Crear workbook con openpyxl para mayor control
        wb = Workbook()

        # ===== HOJA 1: RESERVAS DETALLADAS =====
        ws_reservas = wb.active
        ws_reservas.title = "Reservas Detalladas"

        # Agregar título
        ws_reservas['A1'] = "📊 SISTEMA DE RESERVAS DE COMIDAS - CASINO"
        ws_reservas['A1'].font = Font(size=16, bold=True, color="1F497D")
        ws_reservas.merge_cells('A1:G1')

        # Agregar subtítulo con fechas
        fecha_texto = f"Período: {fecha_inicio or 'Todo'} - {fecha_fin or 'Todo'}"
        ws_reservas['A2'] = fecha_texto
        ws_reservas['A2'].font = Font(size=12, italic=True)
        ws_reservas.merge_cells('A2:G2')

        # Agregar fecha de generación
        ws_reservas['A3'] = f"Generado el: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}"
        ws_reservas['A3'].font = Font(size=10)
        ws_reservas.merge_cells('A3:G3')

        # Escribir headers
        headers = ["Fecha Comida", "Cédula", "Nombre Completo",
                   "Desayuno", "Almuerzo", "Cena", "Reservado En"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_reservas.cell(row=5, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Escribir datos
        for row_num, row_data in enumerate(df.values, 6):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_reservas.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="center")

                # Formato condicional para comidas
                if col_num in [4, 5, 6]:  # Columnas Desayuno, Almuerzo, Cena
                    if value == "Sí":
                        cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    else:
                        cell.fill = PatternFill(
                            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")

        # Crear tabla
        table_range = f"A5:{get_column_letter(len(headers))}{len(df) + 5}"
        table = Table(displayName="ReservasTable", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws_reservas.add_table(table)

        # Ajustar anchos de columna
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            if header == "Nombre Completo":
                ws_reservas.column_dimensions[column_letter].width = 25
            elif header == "Reservado En":
                ws_reservas.column_dimensions[column_letter].width = 18
            else:
                ws_reservas.column_dimensions[column_letter].width = 15

        # ===== HOJA 2: ESTADÍSTICAS =====
        ws_stats = wb.create_sheet("Estadísticas")

        estadisticas = obtener_estadisticas(fecha_inicio, fecha_fin)
        if estadisticas:
            df_stats = pd.DataFrame(estadisticas, columns=[
                "Fecha", "Desayunos", "Almuerzos", "Cenas", "Total Empleados"
            ])

            # Título
            ws_stats['A1'] = "📈 ESTADÍSTICAS DE RESERVAS"
            ws_stats['A1'].font = Font(size=16, bold=True, color="1F497D")
            ws_stats.merge_cells('A1:E1')

            # Headers
            stat_headers = ["Fecha", "Desayunos",
                            "Almuerzos", "Cenas", "Total Empleados"]
            for col_num, header in enumerate(stat_headers, 1):
                cell = ws_stats.cell(row=3, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="9BBB59", end_color="9BBB59", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Datos
            for row_num, row_data in enumerate(df_stats.values, 4):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws_stats.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="center")

            # Crear tabla
            table_range_stats = f"A3:{get_column_letter(len(stat_headers))}{len(df_stats) + 3}"
            table_stats = Table(
                displayName="EstadisticasTable", ref=table_range_stats)
            table_stats.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium3", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            ws_stats.add_table(table_stats)

            # Ajustar anchos
            for col_num in range(1, len(stat_headers) + 1):
                ws_stats.column_dimensions[get_column_letter(
                    col_num)].width = 15

            # ===== GRÁFICO DE BARRAS =====
            # Gráfico de barras para comidas por fecha
            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.style = 10
            bar_chart.title = "Reservas por Tipo de Comida"
            bar_chart.y_axis.title = 'Cantidad'
            bar_chart.x_axis.title = 'Fecha'

            # Datos para el gráfico
            data = Reference(ws_stats, min_col=2, min_row=3,
                             max_col=4, max_row=len(df_stats) + 3)
            cats = Reference(ws_stats, min_col=1, min_row=4,
                             max_row=len(df_stats) + 3)

            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(cats)

            # Posicionar gráfico
            ws_stats.add_chart(bar_chart, "G3")

            # ===== GRÁFICO CIRCULAR =====
            # Gráfico circular para distribución total
            pie_chart = PieChart()
            pie_chart.title = "Distribución Total de Comidas"

            # Calcular totales
            total_desayunos = df_stats["Desayunos"].sum()
            total_almuerzos = df_stats["Almuerzos"].sum()
            total_cenas = df_stats["Cenas"].sum()

            # Crear datos para el pie chart
            pie_data = [
                ['Desayunos', total_desayunos],
                ['Almuerzos', total_almuerzos],
                ['Cenas', total_cenas]
            ]

            # Agregar datos al pie chart
            for i, (label, value) in enumerate(pie_data):
                ws_stats.cell(row=len(df_stats) + 6 +
                              i, column=7).value = label
                ws_stats.cell(row=len(df_stats) + 6 +
                              i, column=8).value = value

            pie_data_ref = Reference(ws_stats, min_col=8, min_row=len(df_stats) + 6,
                                     max_row=len(df_stats) + 8)
            pie_cats_ref = Reference(ws_stats, min_col=7, min_row=len(df_stats) + 6,
                                     max_row=len(df_stats) + 8)

            pie_chart.add_data(pie_data_ref, titles_from_data=False)
            pie_chart.set_categories(pie_cats_ref)

            # Posicionar gráfico circular
            ws_stats.add_chart(pie_chart, "G20")

            # ===== MÉTRICAS RESUMEN =====
            ws_stats['A10'] = "📊 MÉTRICAS RESUMEN"
            ws_stats['A10'].font = Font(size=14, bold=True, color="1F497D")
            ws_stats.merge_cells('A10:E10')

            # Calcular métricas
            total_comidas = total_desayunos + total_almuerzos + total_cenas
            total_empleados = df_stats["Total Empleados"].sum()
            promedio_comidas_dia = total_comidas / \
                len(df_stats) if len(df_stats) > 0 else 0

            metrics = [
                ("Total de Comidas Reservadas", total_comidas),
                ("Total de Empleados", total_empleados),
                ("Promedio de Comidas por Día", round(promedio_comidas_dia, 1)),
                ("Días Analizados", len(df_stats))
            ]

            for i, (label, value) in enumerate(metrics, 11):
                ws_stats.cell(row=i, column=1).value = label
                ws_stats.cell(row=i, column=1).font = Font(bold=True)
                ws_stats.cell(row=i, column=2).value = value
                ws_stats.cell(row=i, column=2).font = Font(
                    bold=True, color="1F497D")

        # ===== HOJA 3: HISTÓRICO MENSUAL =====
        ws_history = wb.create_sheet("Histórico Mensual")

        # Agregar hojas de histórico mensual
        df_history = df.copy()
        df_history['Fecha Comida'] = pd.to_datetime(df_history['Fecha Comida'])
        df_history['Mes'] = df_history['Fecha Comida'].dt.to_period(
            'M').astype(str)

        # Convertir valores "Sí"/"No" a números para poder hacer sum()
        df_history['Desayuno_Num'] = df_history['Desayuno'].map(
            {'Sí': 1, 'No': 0})
        df_history['Almuerzo_Num'] = df_history['Almuerzo'].map(
            {'Sí': 1, 'No': 0})
        df_history['Cena_Num'] = df_history['Cena'].map({'Sí': 1, 'No': 0})

        monthly = df_history.groupby('Mes').agg(
            Desayunos=('Desayuno_Num', 'sum'),
            Almuerzos=('Almuerzo_Num', 'sum'),
            Cenas=('Cena_Num', 'sum'),
            Registros=('Mes', 'count'),
            EmpleadosDistintos=('Cédula', 'nunique')
        ).reset_index()

        # Asegurar que los valores sean numéricos
        monthly['Desayunos'] = pd.to_numeric(
            monthly['Desayunos'], errors='coerce').fillna(0).astype(int)
        monthly['Almuerzos'] = pd.to_numeric(
            monthly['Almuerzos'], errors='coerce').fillna(0).astype(int)
        monthly['Cenas'] = pd.to_numeric(
            monthly['Cenas'], errors='coerce').fillna(0).astype(int)
        monthly['EmpleadosDistintos'] = pd.to_numeric(
            monthly['EmpleadosDistintos'], errors='coerce').fillna(0).astype(int)

        monthly['TotalComidas'] = monthly['Desayunos'] + \
            monthly['Almuerzos'] + monthly['Cenas']
        monthly['PromedioPorEmpleado'] = monthly.apply(
            lambda row: round(row['TotalComidas'] / row['EmpleadosDistintos'],
                              2) if row['EmpleadosDistintos'] > 0 else 0,
            axis=1
        )

        # Escribir título y tabla
        ws_history['A1'] = "📅 HISTÓRICO MENSUAL DE RESERVAS"
        ws_history['A1'].font = Font(size=16, bold=True, color="1F497D")
        ws_history.merge_cells('A1:G1')

        hist_headers = ['Mes', 'Desayunos', 'Almuerzos', 'Cenas', 'Total Comidas',
                        'Registros', 'Empleados Distintos', 'Promedio/Empleado']
        for col_num, header in enumerate(hist_headers, 1):
            cell = ws_history.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_num, row_data in enumerate(monthly[['Mes', 'Desayunos', 'Almuerzos', 'Cenas', 'TotalComidas', 'Registros', 'EmpleadosDistintos', 'PromedioPorEmpleado']].values, 4):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_history.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="center")

        for col_num in range(1, len(hist_headers) + 1):
            ws_history.column_dimensions[get_column_letter(col_num)].width = 18

        # Gráfico de barras mensual en Histórico
        chart_hist = BarChart()
        chart_hist.type = "col"
        chart_hist.style = 12
        chart_hist.title = "Total comidas por mes"
        chart_hist.y_axis.title = 'Cantidad'
        chart_hist.x_axis.title = 'Mes'

        data_hist = Reference(ws_history, min_col=2,
                              min_row=3, max_col=4, max_row=len(monthly) + 3)
        cats_hist = Reference(ws_history, min_col=1,
                              min_row=4, max_row=len(monthly) + 3)
        chart_hist.add_data(data_hist, titles_from_data=True)
        chart_hist.set_categories(cats_hist)

        ws_history.add_chart(chart_hist, "J3")

        # ===== HOJA 4: ESTADÍSTICAS POR EMPLEADO =====
        ws_empleados = wb.create_sheet("Estadísticas por Empleado")

        # Procesar estadísticas por empleado
        df_empleados = df.copy()
        df_empleados['Fecha Comida'] = pd.to_datetime(
            df_empleados['Fecha Comida'])
        df_empleados['Mes'] = df_empleados['Fecha Comida'].dt.to_period(
            'M').astype(str)

        # Convertir valores Sí/No a números para cálculos
        df_empleados['Desayuno_Num'] = df_empleados['Desayuno'].map(
            {'Sí': 1, 'No': 0})
        df_empleados['Almuerzo_Num'] = df_empleados['Almuerzo'].map(
            {'Sí': 1, 'No': 0})
        df_empleados['Cena_Num'] = df_empleados['Cena'].map({'Sí': 1, 'No': 0})

        # Agrupar por empleado y mes
        empleados_stats = df_empleados.groupby(['Cédula', 'Nombre Completo', 'Mes']).agg(
            Desayunos=('Desayuno_Num', 'sum'),
            Almuerzos=('Almuerzo_Num', 'sum'),
            Cenas=('Cena_Num', 'sum'),
            Dias_Registrados=('Fecha Comida', 'nunique')
        ).reset_index()

        # Calcular totales y estadísticas adicionales
        empleados_stats['Total_Comidas'] = empleados_stats['Desayunos'] + \
            empleados_stats['Almuerzos'] + empleados_stats['Cenas']

        # Determinar mayor demanda por empleado
        def get_mayor_demanda(row):
            comidas = {
                'Desayuno': row['Desayunos'],
                'Almuerzo': row['Almuerzos'],
                'Cena': row['Cenas']
            }
            max_comida = max(comidas, key=comidas.get)
            max_cantidad = comidas[max_comida]
            return f"{max_comida} ({max_cantidad})" if max_cantidad > 0 else "Ninguna"

        empleados_stats['Mayor_Demanda'] = empleados_stats.apply(
            get_mayor_demanda, axis=1)

        # Calcular promedio diario
        empleados_stats['Promedio_Diario'] = empleados_stats.apply(
            lambda row: round(row['Total_Comidas'] / row['Dias_Registrados'],
                              2) if row['Dias_Registrados'] > 0 else 0,
            axis=1
        )

        # Escribir título
        ws_empleados['A1'] = "👥 ESTADÍSTICAS POR EMPLEADO"
        ws_empleados['A1'].font = Font(size=16, bold=True, color="1F497D")
        ws_empleados.merge_cells('A1:H1')

        # Headers
        emp_headers = ['Cédula', 'Nombre Completo', 'Mes', 'Desayunos',
                       'Almuerzos', 'Cenas', 'Total Comidas', 'Mayor Demanda', 'Promedio Diario']
        for col_num, header in enumerate(emp_headers, 1):
            cell = ws_empleados.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="548235", end_color="548235", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Datos
        for row_num, row_data in enumerate(empleados_stats[['Cédula', 'Nombre Completo', 'Mes', 'Desayunos', 'Almuerzos', 'Cenas', 'Total_Comidas', 'Mayor_Demanda', 'Promedio_Diario']].values, 4):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_empleados.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="center")

        # Ajustar anchos de columna
        ws_empleados.column_dimensions['A'].width = 15  # Cédula
        ws_empleados.column_dimensions['B'].width = 25  # Nombre Completo
        ws_empleados.column_dimensions['C'].width = 12  # Mes
        ws_empleados.column_dimensions['D'].width = 12  # Desayunos
        ws_empleados.column_dimensions['E'].width = 12  # Almuerzos
        ws_empleados.column_dimensions['F'].width = 8   # Cenas
        ws_empleados.column_dimensions['G'].width = 15  # Total Comidas
        ws_empleados.column_dimensions['H'].width = 18  # Mayor Demanda
        ws_empleados.column_dimensions['I'].width = 16  # Promedio Diario

        # Agregar resumen al final
        resumen_row = len(empleados_stats) + 6
        ws_empleados.cell(
            row=resumen_row, column=1).value = "📊 RESUMEN GENERAL:"
        ws_empleados.cell(row=resumen_row, column=1).font = Font(
            bold=True, size=12)

        # Calcular promedios generales
        if not empleados_stats.empty:
            promedio_general_desayunos = empleados_stats['Desayunos'].mean()
            promedio_general_almuerzos = empleados_stats['Almuerzos'].mean()
            promedio_general_cenas = empleados_stats['Cenas'].mean()
            promedio_general_total = empleados_stats['Total_Comidas'].mean()

            resumen_data = [
                ("Promedio Desayunos por Empleado",
                 round(promedio_general_desayunos, 1)),
                ("Promedio Almuerzos por Empleado",
                 round(promedio_general_almuerzos, 1)),
                ("Promedio Cenas por Empleado", round(promedio_general_cenas, 1)),
                ("Promedio Total Comidas por Empleado",
                 round(promedio_general_total, 1))
            ]

            for i, (label, value) in enumerate(resumen_data, resumen_row + 1):
                ws_empleados.cell(row=i, column=1).value = label
                ws_empleados.cell(row=i, column=1).font = Font(bold=True)
                ws_empleados.cell(row=i, column=2).value = value

        # ===== HOJA 5: DASHBOARD =====
        ws_dashboard = wb.create_sheet("Dashboard")

        # Título del dashboard
        ws_dashboard['A1'] = "🎯 DASHBOARD EJECUTIVO"
        ws_dashboard['A1'].font = Font(size=18, bold=True, color="1F497D")
        ws_dashboard.merge_cells('A1:I1')

        # Información general
        ws_dashboard['A3'] = "Información General"
        ws_dashboard['A3'].font = Font(size=14, bold=True)

        info_general = [
            ("Fecha de Reporte", datetime.now().strftime("%Y/%m/%d %H:%M")),
            ("Período Analizado",
             f"{fecha_inicio or 'Todo'} - {fecha_fin or 'Todo'}"),
            ("Total de Registros", len(df)),
            ("Total de Empleados Únicos", df['Cédula'].nunique())
        ]

        for i, (label, value) in enumerate(info_general, 4):
            ws_dashboard.cell(row=i, column=1).value = label
            ws_dashboard.cell(row=i, column=1).font = Font(bold=True)
            ws_dashboard.cell(row=i, column=2).value = value

        # Estadísticas rápidas
        ws_dashboard['A10'] = "📈 Estadísticas Rápidas"
        ws_dashboard['A10'].font = Font(size=14, bold=True)

        if estadisticas:
            desayuno_total = df_stats["Desayunos"].sum()
            almuerzo_total = df_stats["Almuerzos"].sum()
            cena_total = df_stats["Cenas"].sum()

            stats_rapidas = [
                ("Total Desayunos", desayuno_total),
                ("Total Almuerzos", almuerzo_total),
                ("Total Cenas", cena_total),
                ("Mayor demanda", max([("Desayuno", desayuno_total), ("Almuerzo",
                 almuerzo_total), ("Cena", cena_total)], key=lambda x: x[1])[0])
            ]

            for i, (label, value) in enumerate(stats_rapidas, 11):
                ws_dashboard.cell(row=i, column=1).value = label
                ws_dashboard.cell(row=i, column=1).font = Font(bold=True)
                ws_dashboard.cell(row=i, column=2).value = value
                ws_dashboard.cell(row=i, column=2).font = Font(
                    bold=True, color="1F497D")

        # Ajustar anchos de columna en dashboard
        ws_dashboard.column_dimensions['A'].width = 25
        ws_dashboard.column_dimensions['B'].width = 20

        # Guardar workbook en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    except Exception as e:
        st.error(f"Error al generar Excel avanzado: {e}")
        return None


# ====================== INTERFAZ ======================
st.title("🍽️ Sistema de Reservas de Comidas")
st.caption(
    "Reduce desperdicio de alimentos - Reservas solo día anterior antes de las 22:00")

# Login
if "usuario" not in st.session_state:
    st.subheader("🔑 Autenticación")

    if "pantalla" not in st.session_state:
        st.session_state.pantalla = "Iniciar sesión"

    pantalla = st.radio(
        "Selecciona:",
        ["Iniciar sesión", "Registrarse"],
        index=0 if st.session_state.pantalla == "Iniciar sesión" else 1,
        key="radio_pantalla"
    )

    # Actualizar el estado de la pantalla seleccionada
    st.session_state.pantalla = pantalla

    if pantalla == "Iniciar sesión":
        st.write("### Inicia sesión con tu cédula")
        nombre = st.text_input(
            "Número de cédula", value=st.session_state.get("login_cedula", ""), key="login_cedula")
        password = st.text_input(
            "Contraseña", type="password", value=st.session_state.get("login_password", ""), key="login_password")

        if st.button("Entrar", type="primary", key="btn_login"):
            if nombre and password:
                c.execute(
                    "SELECT nombre, password, rol, nombre_completo FROM usuarios WHERE nombre=? AND password=?", (nombre, password))
                resultado = c.fetchone()
                if resultado:
                    st.session_state.usuario = resultado[0]
                    st.session_state.rol = resultado[2]
                    st.session_state.nombre_completo = resultado[3] if resultado[3] else resultado[0]
                    st.success(
                        f"¡Bienvenido {st.session_state.nombre_completo}! 👋")
                    st.rerun()
                else:
                    st.error("❌ Cédula o contraseña incorrectos")
            else:
                st.error("❌ Por favor completa todos los campos")

    else:
        st.write("### Crear nueva cuenta")
        st.info("📝 Ingresa tu información. Tu usuario será tu número de cédula y la contraseña serán los últimos 4 dígitos.")

        nombre_completo = st.text_input(
            "Nombre completo", value=st.session_state.get("registro_nombre", ""), key="registro_nombre")
        cedula = st.text_input(
            "Número de cédula", value=st.session_state.get("registro_cedula", ""), key="registro_cedula", placeholder="Ejemplo: 1234567890")

        if st.button("Registrarse", type="primary", key="btn_registro"):
            if nombre_completo and cedula:
                resultado = registrar_usuario(nombre_completo, cedula)

                if resultado and resultado != False:
                    _, cedula_user, password_user = resultado
                    st.success("✅ ¡Registro exitoso!")
                    st.info(f"📱 **Usuario:** {cedula_user}")
                    st.info(f"🔐 **Contraseña:** {password_user}")

                    # Cambiar a iniciar sesión y hacer rerun para limpiar campos
                    st.session_state.pantalla = "Iniciar sesión"
                    st.rerun()
            else:
                st.error("❌ Por favor completa todos los campos")
else:
    usuario = st.session_state.usuario
    rol = st.session_state.rol

    # Página especial para admin
    if rol == "admin":
        mostrar_pagina_admin()
    else:
        # Página para empleados
        mostrar_pagina_empleado()

# Footer
st.sidebar.markdown("---")
st.sidebar.caption("App hecha 100% en Python • Streamlit + SQLite")
